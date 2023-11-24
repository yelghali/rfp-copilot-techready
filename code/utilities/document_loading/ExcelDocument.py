from typing import List
from io import BytesIO
from docx import Document
import requests
from .DocumentLoadingBase import DocumentLoadingBase
from ..common.SourceDocument import SourceDocument

import json
import openpyxl
from openpyxl import load_workbook

class ExcelDocumentLoading:
    def __init__(self, file_path):
        self.file_path = file_path

    def process_excel(self):
        workbook = load_workbook(filename=self.file_path)
        excel_data = {}

        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            sheet_data = {}

            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        cell_coordinate = cell.coordinate
                        sheet_data[cell_coordinate] = cell.value

            excel_data[sheet] = sheet_data

        return excel_data

    def to_json(self):
        excel_data = self.process_excel()
        return json.dumps(excel_data, ensure_ascii=False)

# Usage
excel_doc = ExcelDocumentLoading("your_file_path.xlsx")
json_output = excel_doc.to_json()
print(json_output)
